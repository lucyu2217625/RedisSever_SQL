from flask import Flask, render_template, request, Response, redirect, url_for, flash, session,jsonify,render_template_string, send_from_directory
import redis
import json
import os
import redisConnect
import pgConnect
import psycopg2.extras
import threading
from redis.exceptions import RedisError, ConnectionError
from functools import wraps
from werkzeug.security import check_password_hash,generate_password_hash
import time
from datetime import datetime, timezone, timedelta
from collections import defaultdict
import requests
import uuid
import io

# openpyxl 用來解析 Excel (.xlsx) 匯入檔案。
# 如果環境還沒安裝這個套件（requirements.txt 需要加上 openpyxl 並重新 build image），
# 匯入功能裡 Excel 上傳的部分會顯示清楚的錯誤訊息，但 JSON 上傳完全不受影響。
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

def safe_parse_datetime(dt_str):
    """ㄇ
    自定義日期解析，取代 dateutil.parser.parse 以減少外部依賴。
    """
    if not dt_str:
        return None
    dt_str = str(dt_str).strip()
    
    # 處理 ISO 格式 (例如: 2023-10-27T10:30:00)
    try:
        # datetime.fromisoformat 在 3.7+ 支援
        # 替換 Z 為 +00:00 以相容舊版 fromisoformat
        return datetime.fromisoformat(dt_str.replace('Z', '+00:00'))
    except (ValueError, AttributeError):
        pass
    
    # 嘗試其他常見格式
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(dt_str, fmt)
        except ValueError:
            continue
    return None


redisConnect.connect_to_master()
threading.Thread(target=redisConnect.listen_for_failover, daemon=True).start()

pgConnect.connect_to_pg()


def init_core_schema():
    """啟動時檢查系統核心資料表是否存在，沒有就自動建立，有的話直接沿用
    （不會動到既有資料）。統一做法：所有建表一律用 CREATE TABLE IF NOT EXISTS，
    不再依賴 schema.sql 只在 PostgreSQL volume「第一次啟動」時才生效的機制
    ——這樣不管換到哪個環境部署、volume 是不是全新的，這些表都保證存在。

    涵蓋的表：
      users             帳號登入用（username, password_hash, role）
      equipments        機台對照表，eqpid 為主鍵
      contacts          聯絡人對照表，id 為 SERIAL 主鍵
      equipment_history 設備異動紀錄，由 trigger 自動寫入
                        （created/updated/deleted，changed_by 靠
                        set_config('app.current_user', ...) 取得）
      task_history      dispatched_log/failed_queue 的長期歷史存放處，
                        由 history_sync 服務持續從 Redis 搬過來寫入；
                        payload 用 JSONB，因為程式碼裡有用
                        payload->>'EQPID' 這種 JSONB 運算子查詢
    """
    try:
        with pgConnect.get_conn() as conn:
            with conn.cursor() as cur:
                # ── users ──────────────────────────────────────
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS users (
                        username TEXT PRIMARY KEY,
                        password_hash TEXT NOT NULL,
                        role TEXT NOT NULL DEFAULT 'viewer'
                    )
                """)

                # ── equipments ─────────────────────────────────
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS equipments (
                        eqpid TEXT PRIMARY KEY,
                        eqptype TEXT,
                        testerip TEXT,
                        proberip TEXT,
                        linegroup TEXT,
                        floor TEXT,
                        action TEXT,
                        updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
                    )
                """)

                # ── contacts ───────────────────────────────────
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS contacts (
                        id SERIAL PRIMARY KEY,
                        eqpid TEXT,
                        eqptype TEXT,
                        action TEXT,
                        linegroup TEXT,
                        floor TEXT
                    )
                """)

                # ── equipment_history（trigger 記錄 equipments 的異動）──
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS equipment_history (
                        id BIGSERIAL PRIMARY KEY,
                        eqpid TEXT,
                        change_type TEXT NOT NULL,
                        changed_by TEXT,
                        old_value TEXT,
                        new_value TEXT,
                        changed_at TIMESTAMPTZ NOT NULL DEFAULT now()
                    )
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_equipment_history_eqpid_time
                    ON equipment_history (eqpid, changed_at)
                """)
                cur.execute("""
                    CREATE OR REPLACE FUNCTION log_equipment_change()
                    RETURNS TRIGGER AS $$
                    DECLARE
                        v_user TEXT;
                    BEGIN
                        v_user := current_setting('app.current_user', true);
                        IF (TG_OP = 'INSERT') THEN
                            INSERT INTO equipment_history (eqpid, change_type, changed_by, old_value, new_value)
                            VALUES (NEW.eqpid, 'created', v_user, NULL, row_to_json(NEW)::text);
                            RETURN NEW;
                        ELSIF (TG_OP = 'UPDATE') THEN
                            INSERT INTO equipment_history (eqpid, change_type, changed_by, old_value, new_value)
                            VALUES (NEW.eqpid, 'updated', v_user, row_to_json(OLD)::text, row_to_json(NEW)::text);
                            RETURN NEW;
                        ELSIF (TG_OP = 'DELETE') THEN
                            INSERT INTO equipment_history (eqpid, change_type, changed_by, old_value, new_value)
                            VALUES (OLD.eqpid, 'deleted', v_user, row_to_json(OLD)::text, NULL);
                            RETURN OLD;
                        END IF;
                        RETURN NULL;
                    END;
                    $$ LANGUAGE plpgsql;
                """)
                cur.execute("DROP TRIGGER IF EXISTS trg_equipment_history ON equipments")
                cur.execute("""
                    CREATE TRIGGER trg_equipment_history
                    AFTER INSERT OR UPDATE OR DELETE ON equipments
                    FOR EACH ROW EXECUTE FUNCTION log_equipment_change()
                """)

                # ── task_history（history_sync 服務持續寫入）──────
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS task_history (
                        id BIGSERIAL PRIMARY KEY,
                        category TEXT NOT NULL,
                        task_id TEXT,
                        status TEXT NOT NULL,
                        rpa_worker TEXT,
                        dispatch_time TIMESTAMPTZ,
                        payload JSONB
                    )
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_task_history_category_status_time
                    ON task_history (category, status, dispatch_time DESC)
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_task_history_payload_gin
                    ON task_history USING GIN (payload)
                """)
        print("核心資料表（users/equipments/contacts/equipment_history/task_history）已確認存在（不存在則已自動建立）。")
    except Exception as e:
        print(f"Error initializing core schema: {e}")

init_core_schema()


def init_admin_user():
    try:
        with pgConnect.get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM users WHERE username = %s", ('admin',))
                if cur.fetchone() is None:
                    hashed = generate_password_hash('admin123')
                    cur.execute(
                        "INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s)",
                        ('admin', hashed, 'admin')
                    )
                    print("Default admin user created.")
    except Exception as e:
        print(f"Error initializing admin user: {e}")

init_admin_user()


def init_worker_status_history_schema():
    """啟動時檢查 worker_status_history 表是否存在，沒有就自動建立，
    有的話直接沿用（不會動到既有資料）。這樣就不需要另外執行 SQL 腳本，
    也不需要另外的 migration 步驟——跟 init_admin_user() 的做法一致。

    這張表用來記錄 RPA worker（例如 RPA_OP02）的 idle/busy/offline 狀態
    「每一次變化」的時間點，讓 /worker_utilization 可以回推任意時間區間內
    的稼動率（busy 狀態佔區間的比例）。寫入邏輯見 worker_status_sync_loop()。
    """
    try:
        with pgConnect.get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS worker_status_history (
                        id BIGSERIAL PRIMARY KEY,
                        queue_name TEXT NOT NULL,
                        worker_name TEXT NOT NULL,
                        status TEXT NOT NULL,
                        changed_at TIMESTAMPTZ NOT NULL DEFAULT now()
                    )
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_worker_status_history_worker_time
                    ON worker_status_history (worker_name, changed_at)
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_worker_status_history_queue_time
                    ON worker_status_history (queue_name, changed_at)
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_worker_status_history_queue_worker_time
                    ON worker_status_history (queue_name, worker_name, changed_at DESC)
                """)
        print("worker_status_history 表已確認存在（不存在則已自動建立）。")
    except Exception as e:
        print(f"Error initializing worker_status_history schema: {e}")

init_worker_status_history_schema()


def init_contacts_history_schema():
    """啟動時檢查 contacts_history 表與對應的 PostgreSQL trigger 是否存在，
    沒有就自動建立，有的話直接沿用（不會動到既有資料）。做法比照
    init_worker_status_history_schema() / init_admin_user()，不需要另外
    手動執行 SQL 腳本。

    change_type 值：created / updated / deleted
    changed_by：從 set_config('app.current_user', ...) 取得，
                需要在同一個 transaction 裡先呼叫
                pgConnect.set_current_user(conn, username) 才會生效
                （update_contacts() 已經加上這個呼叫）。
    old_value / new_value：用 row_to_json(...)::text 轉成純文字存放。
    """
    try:
        with pgConnect.get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS contacts_history (
                        id BIGSERIAL PRIMARY KEY,
                        contact_id INTEGER,
                        eqpid TEXT,
                        change_type TEXT NOT NULL,
                        changed_by TEXT,
                        old_value TEXT,
                        new_value TEXT,
                        changed_at TIMESTAMPTZ NOT NULL DEFAULT now()
                    )
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_contacts_history_eqpid_time
                    ON contacts_history (eqpid, changed_at)
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_contacts_history_time
                    ON contacts_history (changed_at)
                """)

                cur.execute("""
                    CREATE OR REPLACE FUNCTION log_contacts_change()
                    RETURNS TRIGGER AS $$
                    DECLARE
                        v_user TEXT;
                    BEGIN
                        v_user := current_setting('app.current_user', true);
                        IF (TG_OP = 'INSERT') THEN
                            INSERT INTO contacts_history (contact_id, eqpid, change_type, changed_by, old_value, new_value)
                            VALUES (NEW.id, NEW.eqpid, 'created', v_user, NULL, row_to_json(NEW)::text);
                            RETURN NEW;
                        ELSIF (TG_OP = 'UPDATE') THEN
                            INSERT INTO contacts_history (contact_id, eqpid, change_type, changed_by, old_value, new_value)
                            VALUES (NEW.id, NEW.eqpid, 'updated', v_user, row_to_json(OLD)::text, row_to_json(NEW)::text);
                            RETURN NEW;
                        ELSIF (TG_OP = 'DELETE') THEN
                            INSERT INTO contacts_history (contact_id, eqpid, change_type, changed_by, old_value, new_value)
                            VALUES (OLD.id, OLD.eqpid, 'deleted', v_user, row_to_json(OLD)::text, NULL);
                            RETURN OLD;
                        END IF;
                        RETURN NULL;
                    END;
                    $$ LANGUAGE plpgsql;
                """)
                cur.execute("""
                    DROP TRIGGER IF EXISTS trg_contacts_history ON contacts
                """)
                cur.execute("""
                    CREATE TRIGGER trg_contacts_history
                    AFTER INSERT OR UPDATE OR DELETE ON contacts
                    FOR EACH ROW EXECUTE FUNCTION log_contacts_change()
                """)
        print("contacts_history 表與 trigger 已確認存在（不存在則已自動建立）。")
    except Exception as e:
        print(f"Error initializing contacts_history schema: {e}")

init_contacts_history_schema()


# ── Worker 狀態歷史背景同步 ──────────────────────────────────────
# 跟 redisConnect.listen_for_failover 一樣，用背景 daemon thread 常駐執行，
# 不需要另外開一個容器/服務。持續輪詢 Redis 的 worker_status Hash，
# 偵測到某個 worker 的 idle/busy/offline 狀態變化時，才寫一筆紀錄進
# worker_status_history（狀態沒變化不會重複寫入，避免資料量爆炸）。
#
# ⚠️ 限制：只能從這個背景執行緒開始運行之後才有歷史資料，
# 之前的 busy/idle 變化沒有被記錄下來。
WORKER_STATUS_QUEUES_FOR_SYNC = [
    "LotActions_worker_status",
    "prober_worker_status",
    "LineNotify_worker_status",
]
WORKER_STATUS_SYNC_INTERVAL_SEC = 3  # 輪詢間隔（秒）

# 記憶體快取：{(queue_name, worker_name): last_status}，用來判斷狀態是否真的變化
_worker_status_last_known = {}


def _load_worker_status_initial_state():
    """啟動時，從 PostgreSQL 撈每個 worker 最後一筆紀錄的狀態，當作起始基準，
    避免程式重啟後把「其實沒有變化」誤判成「變化」而重複寫入。"""
    try:
        with pgConnect.get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT DISTINCT ON (queue_name, worker_name)
                        queue_name, worker_name, status
                    FROM worker_status_history
                    ORDER BY queue_name, worker_name, changed_at DESC
                """)
                for queue_name, worker_name, status in cur.fetchall():
                    _worker_status_last_known[(queue_name, worker_name)] = status
        print(f"[worker_status_sync] 已載入 {len(_worker_status_last_known)} 個 worker 的最後狀態")
    except Exception as e:
        print(f"[worker_status_sync] 載入初始狀態失敗：{e}")


def _worker_status_extract_status(raw_val):
    """worker_status Hash 的 value 是 JSON 字串，例如 {"status": "busy", ...}。
    解析邏輯跟 worker_status_partial() 保持一致。"""
    try:
        data = json.loads(raw_val)
        return str(data.get("status", "unknown")).strip().lower()
    except Exception:
        return "unknown"


def _worker_status_sync_once():
    """跑一輪：檢查三個 worker_status Hash 裡的每個 worker，
    狀態跟上次記錄的不一樣就寫入一筆新紀錄。"""
    for queue_name in WORKER_STATUS_QUEUES_FOR_SYNC:
        try:
            all_workers = redisConnect.redis_master.hgetall(queue_name)
        except Exception as e:
            print(f"[worker_status_sync] 讀取 {queue_name} 失敗：{e}")
            continue

        for worker_name, raw_val in all_workers.items():
            status = _worker_status_extract_status(raw_val)
            key = (queue_name, worker_name)
            prev_status = _worker_status_last_known.get(key)

            if prev_status == status:
                continue  # 狀態沒變化，不寫入

            try:
                with pgConnect.get_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO worker_status_history (queue_name, worker_name, status)
                            VALUES (%s, %s, %s)
                            """,
                            (queue_name, worker_name, status)
                        )
                _worker_status_last_known[key] = status
                print(f"[worker_status_sync] {queue_name}/{worker_name}: "
                      f"{prev_status or '(初次記錄)'} -> {status}")
            except Exception as e:
                print(f"[worker_status_sync] 寫入失敗 {queue_name}/{worker_name}：{e}")


def worker_status_sync_loop():
    """背景常駐迴圈，跟 redisConnect.listen_for_failover 一樣用 daemon thread 啟動。"""
    _load_worker_status_initial_state()
    print("worker 狀態歷史背景同步已啟動，開始持續輪詢…")
    while True:
        try:
            _worker_status_sync_once()
        except Exception as e:
            print(f"⚠️ [worker_status_sync] 同步發生錯誤，{WORKER_STATUS_SYNC_INTERVAL_SEC} 秒後重試：{e}")
        time.sleep(WORKER_STATUS_SYNC_INTERVAL_SEC)


threading.Thread(target=worker_status_sync_loop, daemon=True).start()

app = Flask(__name__, static_folder="static")
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'your-secret-key')  # 請設置安全的密鑰
app.config['SESSION_PERMANENT'] = False  # 設置 session 為非永久，瀏覽器關閉後失效

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(app.static_folder, 'favicon.ico', mimetype='image/vnd.microsoft.icon')

# 預定義可查詢的佇列名稱
ALLOWED_QUEUES = os.getenv('ALLOWED_QUEUES', 'processing_queue,failed_queue,task_queue,retry_queue,worker_status,dispatched_log').split(',')

# Dashboard「Queue 狀態」面板專用的精簡監控清單（畫面空間有限，只挑重點 Queue）。
# 注意：/queue_lengths、/queue_lengths_partial 顯示「全部」ALLOWED_QUEUES，
# 範圍比這份清單大，詳見 queue_lengths() 的說明。
DASHBOARD_MONITORED_QUEUES = [
    "LotActions_task_queue",
    "LotActions_failed_queue",
    "LotActions_dispatched_log",
    "LotActions_processing_queue",
    "LineNotify_task_queue",
    "LineNotify_failed_queue",
    "prober_task_queue",
    "prober_failed_queue",
]

# 凡是以 "_dispatched_log" 結尾的 Queue（LotActions_dispatched_log、
# prober_dispatched_log、LineNotify_dispatched_log...），長度一律改成
# 查詢 PostgreSQL task_history 的累計總筆數，而不是 Redis LLEN。
# 原因：sync_history_to_postgres.py 會持續把 Redis 裡的資料搬到
# PostgreSQL 後清空（LMOVE + LTRIM），所以 Redis LLEN 只能反映
# 「還沒被同步搬走的暫存量」（通常接近 0），無法代表真正的總筆數。
# 用 endswith 判斷而非寫死清單，未來新增其他 category 的 dispatched_log
# 不需要再改程式碼。
PG_BACKED_SUFFIXES = (
    "_dispatched_log",
)


def is_pg_backed_queue(q):
    return q.endswith(PG_BACKED_SUFFIXES)


def get_pg_task_count(queue_name):
    """把 queue_name（例如 LotActions_dispatched_log）比照 history_queue_spec()
    的規則拆成 (category, status)，查詢 PostgreSQL task_history 表的累計總筆數。"""
    category, status = history_queue_spec(queue_name)
    if not category:
        return 0
    try:
        with pgConnect.get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT count(*) FROM task_history WHERE category = %s AND status = %s",
                    (category, status)
                )
                row = cur.fetchone()
                return row[0] if row else 0
    except Exception:
        return 0


def get_queue_length(q):
    """回傳單一佇列的 (type, length)。
    - 若 q 以 "_dispatched_log" 結尾：長度改成查 PostgreSQL task_history 的累計總筆數。
    - 其餘情況：沿用原本 Redis TYPE + LLEN 的即時查詢邏輯
      （只有 list 型態才用 LLEN 計算長度，其餘型態一律回傳 0）。
    這個函式同時被 queue_lengths()、queue_lengths_partial()、dashboard_data()
    三處呼叫，確保計算邏輯永遠一致。"""
    if is_pg_backed_queue(q):
        return "list", get_pg_task_count(q)
    try:
        q_type = redisConnect.redis_master.type(q)
        length = redisConnect.redis_master.llen(q) if q_type == "list" else 0
        return q_type, length
    except redis.RedisError:
        return "unknown", 0

# 自定義 JSON 過濾器
@app.template_filter('from_json')
def from_json_filter(s):
    if isinstance(s, str):
        try:
            return json.loads(s)
        except json.JSONDecodeError:
            return {}
    return s

# 注入 max 和 min 函數供模板使用
@app.context_processor
def inject_functions():
    return dict(max=max, min=min)

# 檢查是否登入的裝飾器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            # flash('請先登入', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# 檢查角色權限的裝飾器
def role_required(role):
    @wraps(role)
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'username' not in session:
                # flash('請先登入', 'error')
                return redirect(url_for('login'))
            try:
                with pgConnect.get_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT role FROM users WHERE username = %s", (session['username'],))
                        row = cur.fetchone()
                if not row:
                    flash('使用者資料不存在', 'error')
                    return redirect(url_for('login'))
                user_role = row[0]
                if user_role != role:
                    flash('您沒有權限執行此操作', 'error')
                    return redirect(url_for('index'))
                return f(*args, **kwargs)
            except Exception as e:
                flash('無法連接到資料庫，請稍後再試', 'error')
                return redirect(url_for('login'))
        return decorated_function
    return decorator

# 登入頁面
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        try:
            with pgConnect.get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT password_hash, role FROM users WHERE username = %s",
                        (username,)
                    )
                    row = cur.fetchone()
            if row:
                stored_hash, stored_role = row
                if check_password_hash(stored_hash, password):
                    session['username'] = username
                    session['role'] = stored_role
                    # flash('登入成功！', 'success')
                    return redirect(url_for('index'))
                else:
                    return render_template('login.html', error=f'使用者名稱或密碼錯誤')
            else:
                return render_template('login.html', error='使用者名稱或密碼錯誤')
        except Exception:
            return render_template('login.html', error='無法連接到資料庫，請稍後再試')

    return render_template('login.html')

# 登出
@app.route('/logout')
def logout():
    session.pop('username', None)
    session.pop('role', None)
    # flash('您已登出', 'success')
    return redirect(url_for('login'))

# 主頁（佇列查詢）
@app.route('/', methods=['GET', 'POST'])
@login_required
def index():
    queue_name = ''
    queue_data = []
    queue_length = 0
    error = None
    page = int(request.args.get('page', 1))
    
    # Per page logic
    try:
        per_page_source = request.form if request.method == 'POST' else request.args
        per_page = int(per_page_source.get('per_page', 10))
    except (ValueError, TypeError):
        per_page = 10
    if per_page not in [10, 25, 50, 100]: # Validate allowed values
        per_page = 10
    is_hash = False
    hash_data = {}
    is_history_queue = False
    user_role = session.get('role', 'viewer')
    total_pages = 1 # Initialize total_pages here

    if request.method == 'POST':
        queue_name = request.form.get('queue_name', '').strip()
    else:
        queue_name = request.args.get('queue_name', '').strip()

    available_queues = ALLOWED_QUEUES

    history_view = load_history_queue(queue_name, page, per_page) if queue_name else None
    if history_view:
        queue_data = history_view['queue_data']
        queue_length = history_view['queue_length']
        total_pages = history_view['total_pages'] or 1
        is_history_queue = True
    elif queue_name in ALLOWED_QUEUES:
        try:
            if redisConnect.redis_master.exists(queue_name):
                key_type = redisConnect.type(queue_name)
                if key_type == 'list':
                    total_queue_length = redisConnect.redis_master.llen(queue_name)
                    if total_queue_length > 0:
                        start = (page - 1) * per_page
                        end = start + per_page * 5

                        raw_data = redisConnect.redis_master.lrange(queue_name, 0, -1)

                        filtered_data = []
                        for raw in raw_data:
                            # 確保 raw 是字串格式
                            raw_str = raw.decode() if isinstance(raw, bytes) else raw
                            parsed = parse_json(raw_str)
                            if isinstance(parsed, dict) and parsed.get('task_id') == '__INIT__':
                                continue
                            filtered_data.append({'parsed': parsed, 'raw': raw_str})

                        queue_length = len(filtered_data)
                        total_pages = (queue_length + per_page - 1) // per_page

                        start = (page - 1) * per_page
                        end = start + per_page
                        queue_data = filtered_data[start:end]

                        queue_length = total_queue_length
                        if not queue_data and start == 0:
                            error = f"「{queue_name}」沒有內容。"
                    else:
                        error = f"「{queue_name}」沒有內容。"
                        
                elif key_type == 'hash':
                    is_hash = True
                    raw_hash_data = redisConnect.redis_master.hgetall(queue_name)
                    print(f"[DEBUG] hgetall({queue_name}) -> {raw_hash_data}")   # 🔥 log 1
                    
                    hash_data = {} 
                    if raw_hash_data:
                        for key, val in raw_hash_data.items():
                            key_str = key.decode() if isinstance(key, bytes) else key
                            val_str = val.decode() if isinstance(val, bytes) else val
                            print(f"[DEBUG] key={key_str}, val={val_str}")       # 🔥 log 2
                            try:
                                json_val = json.loads(val_str)
                                hash_data[key_str] = json_val.get("status", "未知")
                            except (json.JSONDecodeError, AttributeError) as e:
                                print(f"[ERROR] JSON parse error: {e} for val={val_str}")
                                hash_data[key_str] = "無法解析"
                    else:
                        error = f"「{queue_name}」沒有內容。"
                        print(f"[DEBUG] {error}")
                else:
                    error = f"「{queue_name}」的資料型態不受支持。"
                    print(f"[DEBUG] {error}")


            else:
                error = f"「{queue_name}」沒有內容。"
        except (RedisError, ConnectionError) as e:
            error = "無法連接到資料庫，請稍後再試。"
            redisConnect.connect_to_master()
    elif queue_name:
        error = "請選擇一個有效的佇列名稱。"



    return render_template(
        'queue_viewer.html',
        queue_name=queue_name,
        queue_data=queue_data,
        queue_length=queue_length,
        error=error,
        available_queues=available_queues,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        is_hash=is_hash,
        hash_data=hash_data,
        is_history_queue=is_history_queue,
        user_role=user_role,
        active_page='index'
    )

def parse_json(item):
    try:
        return json.loads(item)
    except json.JSONDecodeError:
        return item


def history_queue_spec(queue_name):
    for suffix, status in (
        ('_dispatched_log', 'dispatched'),
        ('_failed_queue', 'failed'),
    ):
        if queue_name.endswith(suffix):
            return queue_name[:-len(suffix)], status
    return None, None


def normalize_history_row(row):
    payload = row.get('payload') or {}
    if not isinstance(payload, dict):
        payload = {'payload': payload}

    dispatch_time = row.get('dispatch_time')
    if isinstance(dispatch_time, datetime):
        dispatch_time = dispatch_time.astimezone(timezone(timedelta(hours=8))).strftime('%Y-%m-%d %H:%M:%S')

    display = dict(payload)
    display.update({
        'category': row.get('category'),
        'task_id': row.get('task_id'),
        'status': row.get('status'),
        'rpa_worker': row.get('rpa_worker'),
        'dispatch_time': dispatch_time,
    })

    return {
        'parsed': display,
        'raw': json.dumps(display, ensure_ascii=False),
        # task_history 的主鍵 id，用來讓「刪除所選」可以精準指定要刪哪一列，
        # 不必再靠比對整段 JSON 內容（Redis 版本才需要那樣做）。
        'pg_id': row.get('id'),
    }


def load_history_queue(queue_name, page, per_page):
    category, status = history_queue_spec(queue_name)
    if not category:
        return None

    offset = (page - 1) * per_page
    with pgConnect.get_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT count(*) AS cnt
                FROM task_history
                WHERE category = %s AND status = %s
                """,
                (category, status)
            )
            total = cur.fetchone()['cnt']

            cur.execute(
                """
                SELECT id, category, task_id, status, rpa_worker, dispatch_time, payload
                FROM task_history
                WHERE category = %s AND status = %s
                ORDER BY dispatch_time DESC, id DESC
                LIMIT %s OFFSET %s
                """,
                (category, status, per_page, offset)
            )
            rows = cur.fetchall()

    queue_data = [normalize_history_row(row) for row in rows]
    total_pages = (total + per_page - 1) // per_page if total else 0
    return {
        'queue_data': queue_data,
        'queue_length': total,
        'total_pages': total_pages,
        'is_history_queue': True,
    }


def delete_history_rows(queue_name, pg_ids):
    """從 PostgreSQL task_history 表刪除指定的資料列（依主鍵 id）。
    只會刪除屬於 queue_name 對應 category/status 的資料列，避免誤刪其他佇列的紀錄。"""
    category, status = history_queue_spec(queue_name)
    if not category or not pg_ids:
        return 0
    with pgConnect.get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                DELETE FROM task_history
                WHERE category = %s AND status = %s AND id = ANY(%s)
                """,
                (category, status, pg_ids)
            )
            return cur.rowcount


def clear_history_queue(queue_name):
    """清空 PostgreSQL task_history 表裡屬於 queue_name 對應 category/status 的全部資料列。"""
    category, status = history_queue_spec(queue_name)
    if not category:
        return 0
    with pgConnect.get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM task_history WHERE category = %s AND status = %s",
                (category, status)
            )
            return cur.rowcount

def load_equipments_for_download():
    """比照 equipments() route 的查詢邏輯，回傳可直接 json.dumps 的 dict。
    equipments 資料來源已搬遷到 PostgreSQL，/download 需要跟著查 Postgres，
    不能再假設資料還留在 Redis 的 equipments hash 裡。"""
    data = {}
    with pgConnect.get_conn() as conn:
        with pgConnect.dict_cursor(conn) as cur:
            cur.execute("SELECT * FROM equipments ORDER BY eqpid")
            rows = cur.fetchall()
    for row in rows:
        data[row['eqpid']] = {
            'EQPTYPE': row['eqptype'],
            'TESTERIP': row['testerip'],
            'PROBERIP': row['proberip'],
            'LINEGROUP': row['linegroup'],
            'floor': row['floor'],
            'Action': row['action'],
        }
    return data


def load_contacts_for_download():
    """比照 contacts() route 的查詢邏輯，回傳可直接 json.dumps 的 dict。
    contacts 資料來源已搬遷到 PostgreSQL，/download 需要跟著查 Postgres。"""
    data = {}
    with pgConnect.get_conn() as conn:
        with pgConnect.dict_cursor(conn) as cur:
            cur.execute("SELECT * FROM contacts ORDER BY id")
            rows = cur.fetchall()
    for row in rows:
        data[str(row['id'])] = {
            'EQPID': row['eqpid'],
            'EQPTYPE': row['eqptype'],
            'Action': row['action'],
            'LINEGROUP': row['linegroup'],
            'floor': row['floor'],
        }
    return data


@app.route('/download')
@login_required
def download():
    queue_name = request.args.get('queue_name', '').strip()

    if not queue_name:
        return "錯誤：請選擇有效的佇列名稱", 400

    try:
        data = []
        # equipments / contacts 這兩個資料來源已經搬遷到 PostgreSQL
        # （equipments()/contacts() route 都已改查 Postgres），
        # /download 要跟著改，不能再走 Redis hash 那條路徑。
        if queue_name == 'equipments':
            data = load_equipments_for_download()
            if not data:
                return "「equipments」沒有內容。", 404
        elif queue_name == 'contacts':
            data = load_contacts_for_download()
            if not data:
                return "「contacts」沒有內容。", 404
        else:
            history_view = load_history_queue(queue_name, 1, 1000000) if queue_name else None
            if history_view:
                data = [item['parsed'] for item in history_view['queue_data']]
            elif redisConnect.redis_master.exists(queue_name):
                key_type = redisConnect.type(queue_name)
                if key_type == 'list':
                    # Since decode_responses=True, lrange returns list of strings
                    raw_data = redisConnect.redis_master.lrange(queue_name, 0, -1)
                    for item in raw_data:
                        parsed = parse_json(item)
                        if parsed is not None and not (isinstance(parsed, dict) and parsed.get('task_id') == '__INIT__'):
                            data.append(parsed)
                elif key_type == 'hash':
                    # Since decode_responses=True, hgetall returns dict of str:str
                    hash_data = redisConnect.redis_master.hgetall(queue_name)
                    data = {}
                    for key, val in hash_data.items():
                        # The key is already a string.
                        # The value is a string, which might be JSON.
                        data[key] = parse_json(val)
                else:
                    return "不支援的資料類型", 400
            else:
                return "佇列不存在或為空", 404

        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        return Response(
            json_str,
            mimetype='application/json',
            headers={"Content-Disposition": f"attachment;filename={queue_name}.json"}
        )
    except Exception as e:
        return f"下載失敗：{str(e)}", 500


@app.route('/delete_selected', methods=['POST'])
@login_required
@role_required('admin')
def delete_selected():
    queue_name = request.form.get('queue_name', '').strip()
    delete_all_force = request.form.get('delete_all_force') == '1'
    selected_items = request.form.getlist('selected_items[]')

    if not queue_name:
        flash("缺少佇列名稱", "error")
        return redirect(url_for('index'))

    # 歷史紀錄類 Queue（_dispatched_log / _failed_queue 結尾）資料實際存放在
    # PostgreSQL task_history 表，刪除邏輯改成對 Postgres 下 DELETE，
    # 而不是操作 Redis（Redis 端這幾個 key 通常已經被同步服務搬空，操作也無意義）。
    # 勾選框的 value 會是 task_history 的主鍵 id（見 normalize_history_row 的 pg_id），
    # 「清空全部」則直接刪除該 category+status 底下的全部資料列。
    if history_queue_spec(queue_name)[0]:
        try:
            if delete_all_force:
                count = clear_history_queue(queue_name)
                flash(f'已清空「{queue_name}」的歷史紀錄，共 {count} 筆', 'success')
                return redirect(url_for('index', queue_name=queue_name))

            if not selected_items:
                flash("請選擇要刪除的項目", "error")
                return redirect(url_for('index', queue_name=queue_name))

            try:
                pg_ids = [int(v) for v in selected_items]
            except ValueError:
                flash("刪除失敗：選取的項目格式不正確", "error")
                return redirect(url_for('index', queue_name=queue_name))

            count = delete_history_rows(queue_name, pg_ids)
            flash(f'已成功從「{queue_name}」刪除 {count} 筆歷史紀錄', 'success')
        except Exception as e:
            flash(f"刪除失敗：{str(e)}", "error")

        resp = redirect(url_for('index', queue_name=queue_name))
        resp.headers['Location'] = resp.headers['Location'].replace(':5000', '')
        return resp

    try:
        if delete_all_force:
            redisConnect.redis_master.delete(queue_name)
            flash(f'已清空整個佇列「{queue_name}」', 'success')
            return redirect(url_for('index', queue_name=queue_name))

        if not selected_items:
            flash("請選擇要刪除的項目", "error")
            return redirect(url_for('index', queue_name=queue_name))

        key_type = redisConnect.redis_master.type(queue_name)
        if key_type == 'list':
            count = 0
            # 取得 Redis 裡所有的原始資料
            all_raw = redisConnect.redis_master.lrange(queue_name, 0, -1)
            
            for item_val in selected_items:
                # 1. 嘗試直接刪除（完全匹配）
                if redisConnect.redis_master.lrem(queue_name, 1, item_val):
                    count += 1
                else:
                    # 2. 如果失敗，嘗試解析後再標準化刪除（解決空白字元問題）
                    try:
                        normalized_val = json.dumps(json.loads(item_val), separators=(',', ':'), ensure_ascii=False)
                        # 在 Redis 裡尋找符合標準化格式的項並刪除
                        for raw in all_raw:
                            try:
                                if json.dumps(json.loads(raw), separators=(',', ':'), ensure_ascii=False) == normalized_val:
                                    if redisConnect.redis_master.lrem(queue_name, 1, raw):
                                        count += 1
                                        break
                            except: continue
                    except:
                        pass
            flash(f'已成功從「{queue_name}」刪除 {count} 筆資料', 'success')
        elif key_type == 'hash':
            count = redisConnect.redis_master.hdel(queue_name, *selected_items)
            flash(f'已成功從「{queue_name}」刪除 {count} 個欄位', 'success')
        else:
            flash(f"不支援的資料類型: {key_type}", "error")
    except (RedisError, ConnectionError) as e:
        flash(f"刪除失敗：{str(e)}", "error")
        redisConnect.connect_to_master()

    # 解決 Codespaces 重導向問題：使用相對路徑
    resp = redirect(url_for('index', queue_name=queue_name))
    resp.headers['Location'] = resp.headers['Location'].replace(':5000', '')
    return resp

@app.route('/update_status', methods=['POST'])
@login_required
@role_required('admin')
def update_status():
    queue_name = request.form.get('queue_name')
    if not queue_name:
        return "缺少 queue_name", 400

    key_type = redisConnect.redis_master.type(queue_name)
    if key_type != 'hash':
        return f"佇列 {queue_name} 不是 hash 型態! 是 {key_type}", 400

    # 取得所有 status_* 欄位並轉成 rpa_name → status 的 dict
    updates = {
        key.replace("status_", ""): value
        for key, value in request.form.items()
        if key.startswith("status_")
    }

    try:
        with redisConnect.redis_master.pipeline() as pipe:
            for rpa_name, new_status in updates.items():
                raw = redisConnect.redis_master.hget(queue_name, rpa_name)
                try:
                    data = json.loads(raw)
                    data["status"] = new_status
                except (json.JSONDecodeError, TypeError):
                    data = {"status": new_status}  

                pipe.hset(queue_name, rpa_name, json.dumps(data))
            pipe.execute()
    except Exception as e:
        return f"更新失敗：{str(e)}", 500

    return redirect(url_for("index", queue_name=queue_name))

@app.route("/queue_lengths")
def queue_lengths():
    grouped_queues = {}

    # 顯示「全部」ALLOWED_QUEUES（不只是 dashboard 的 8 個重點 Queue），
    # 讓 LotActions / LineNotify / prober 三個 category 底下的
    # task_queue、failed_queue、processing_queue、retry_queue、dispatched_log
    # 都能完整列出來。worker_status 是 hash 型態，維持原本排除規則
    # （queue_lengths.html 樣板本身也會用 q.type != 'hash' 再過濾一次）。
    # 長度計算一律呼叫共用的 get_queue_length()，確保跟 dashboard 的
    # 計算邏輯一致（例如 _dispatched_log 結尾會改查 PostgreSQL 累計總筆數）。
    for q in ALLOWED_QUEUES:
        if 'worker_status' in q:
            continue
        q_type, length = get_queue_length(q)

        prefix = q.split("_", 1)[0]
        if prefix not in grouped_queues:
            grouped_queues[prefix] = []
        grouped_queues[prefix].append({"name": q, "length": length, "type": q_type})

    return render_template("queue_lengths.html", grouped_queues=grouped_queues, active_page='queue_lengths')

@app.route("/queue_lengths_partial")
def queue_lengths_partial():
    grouped_queues = {}
    for q in ALLOWED_QUEUES:
        if 'worker_status' in q:
            continue
        q_type, length = get_queue_length(q)

        prefix = q.split("_", 1)[0]
        if prefix not in grouped_queues:
            grouped_queues[prefix] = []
        grouped_queues[prefix].append({"name": q, "length": length, "type": q_type})

    tbodies = []
    for prefix, queues in grouped_queues.items():
        tbody_html = render_template_string('''
            {% for q in queues %}
                {% if q.type != 'hash' %}
                <tr class="hover:bg-blue-50 cursor-pointer transition-colors group"
                    onclick="window.location.href='/?queue_name={{ q.name }}'">
                    <td class="py-3 px-4 font-mono text-xs text-gray-600 group-hover:text-blue-700">
                        {{ q.name }}
                    </td>
                    <td class="py-3 px-4 text-right">
                        <span class="{% if q.length > 20 %}text-red-600 font-bold{% else %}text-gray-900{% endif %}">
                            {{ q.length }}
                        </span>
                    </td>
                </tr>
                {% endif %}
            {% endfor %}
        ''', queues=queues)
        tbodies.append(tbody_html)

    return jsonify(tbodies)

@app.route("/worker_status")
def worker_status():
    return render_template("worker_status.html", active_page='worker_status')

@app.route("/worker_status_partial")
def worker_status_partial():
    import json
    queues = ["prober_worker_status", "LineNotify_worker_status", "LotActions_worker_status"]
    grouped_workers = {}

    try:
        for queue_name in queues:
            queue_workers = {}
            all_workers = redisConnect.redis_master.hgetall(queue_name)
            for k, v in all_workers.items():
                key_str = k.decode() if isinstance(k, bytes) else str(k)
                val_str = v.decode() if isinstance(v, bytes) else str(v)

                try:
                    json_val = json.loads(val_str)
                    status = json_val.get("status", "unknown").strip().lower()
                except Exception:
                    status = "unknown"

                queue_workers[key_str] = status
            grouped_workers[queue_name] = queue_workers
    except redis.RedisError:
        grouped_workers = {}

    return jsonify(grouped_workers)



@app.route('/update_equipments', methods=['POST'])
@login_required
def update_equipments():
    edit_mode = request.form.get('edit_mode', '0')
    current_user = session.get('username', '')

    if edit_mode == '1':
        with pgConnect.get_conn() as conn:
            # 讓 equipment_history 的 trigger 記得是誰做的異動
            pgConnect.set_current_user(conn, current_user)

            with conn.cursor() as cur:
                cur.execute("SELECT eqpid, eqptype, testerip, proberip, linegroup, floor, action FROM equipments")
                existing_map = {
                    row[0]: (row[1] or '', row[2] or '', row[3] or '', row[4] or '', row[5] or '', row[6] or '')
                    for row in cur.fetchall()
                }
                existing_ids = list(existing_map.keys())

                for key in existing_ids:
                    if request.form.get(f'delete_{key}'):
                        cur.execute("DELETE FROM equipments WHERE eqpid = %s", (key,))
                        continue

                    eqptype = request.form.get(f'EQPTYPE_{key}', '').strip()
                    testerip = request.form.get(f'TESTERIP_{key}', '').strip()
                    proberip = request.form.get(f'PROBERIP_{key}', '').strip()
                    linegroup = request.form.get(f'LINEGROUP_{key}', '').strip()
                    floor = request.form.get(f'floor_{key}', '').strip()
                    action = request.form.get(f'ACTION_{key}', '').strip()

                    if eqptype and testerip and proberip and linegroup and floor and action:
                        new_vals = (eqptype, testerip, proberip, linegroup, floor, action)
                        old_vals = existing_map.get(key)
                        if old_vals != new_vals:
                            cur.execute(
                                """
                                UPDATE equipments
                                   SET eqptype = %s, testerip = %s, proberip = %s,
                                       linegroup = %s, floor = %s, action = %s,
                                       updated_at = now()
                                 WHERE eqpid = %s
                                """,
                                (eqptype, testerip, proberip, linegroup, floor, action, key)
                            )

                new_eqpids = request.form.getlist('new_EQPID[]')
                new_eqptypes = request.form.getlist('new_EQPTYPE[]')
                new_testerips = request.form.getlist('new_TESTERIP[]')
                new_proberips = request.form.getlist('new_PROBERIP[]')
                new_linegroups = request.form.getlist('new_LINEGROUP[]')
                new_floors = request.form.getlist('new_floor[]')
                new_actions = request.form.getlist('new_ACTION[]')

                for i in range(len(new_eqpids)):
                    eqpid = new_eqpids[i].strip()
                    if not eqpid or eqpid in existing_ids:
                        continue

                    if (new_eqptypes[i].strip() and new_testerips[i].strip() and
                            new_proberips[i].strip() and new_linegroups[i].strip() and
                            new_floors[i].strip() and new_actions[i].strip()):
                        cur.execute(
                            """
                            INSERT INTO equipments (eqpid, eqptype, testerip, proberip, linegroup, floor, action)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                eqpid,
                                new_eqptypes[i].strip(),
                                new_testerips[i].strip(),
                                new_proberips[i].strip(),
                                new_linegroups[i].strip(),
                                new_floors[i].strip(),
                                new_actions[i].strip(),
                            )
                        )
        return redirect(url_for("equipments"))

    return redirect(url_for("equipments"))


@app.route('/equipment_history')
@login_required
def equipment_history():
    """查看單一 EQPID 的異動紀錄（equipment_history 表由 PostgreSQL trigger 自動寫入）"""
    eqpid = request.args.get('eqpid', '').strip()
    history = []
    error = None
    try:
        with pgConnect.get_conn() as conn:
            with pgConnect.dict_cursor(conn) as cur:
                if eqpid:
                    cur.execute(
                        """
                        SELECT * FROM equipment_history
                         WHERE eqpid = %s
                         ORDER BY changed_at DESC
                         LIMIT 200
                        """,
                        (eqpid,)
                    )
                else:
                    cur.execute(
                        """
                        SELECT * FROM equipment_history
                         ORDER BY changed_at DESC
                         LIMIT 200
                        """
                    )
                history = cur.fetchall()
        if not history:
            error = "沒有找到異動紀錄。"
    except Exception as e:
        error = f"查詢失敗：{e}"

    return render_template(
        'equipment_history.html',
        history=history,
        eqpid=eqpid,
        error=error,
        active_page='equipment_history'
    )

@app.route('/register', methods=['GET', 'POST'])
@role_required('admin')
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        try:
            with pgConnect.get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT 1 FROM users WHERE username = %s", (username,))
                    if cur.fetchone():
                        flash('使用者已存在', 'error')
                    else:
                        hashed = generate_password_hash(password)
                        cur.execute(
                            "INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s)",
                            (username, hashed, role)
                        )
                        flash('使用者新增成功', 'success')
        except Exception as e:
            flash(f'新增失敗：{e}', 'error')
        return redirect(url_for('register'))

    # 回傳目前所有使用者
    user_list = []
    with pgConnect.get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT username, role FROM users ORDER BY username")
            user_list = [{'username': u, 'role': r} for u, r in cur.fetchall()]

    return render_template('register.html', user_list=user_list, active_page='register')


@app.route('/delete_user', methods=['POST'])
@role_required('admin')
def delete_user():
    username = request.form['username']
    if username == 'admin':
        flash('不能刪除 admin 使用者', 'error')
    else:
        with pgConnect.get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM users WHERE username = %s", (username,))
        flash(f'{username} 已刪除', 'success')
    return redirect(url_for('register'))


@app.route('/update_user_password', methods=['POST'])
@role_required('admin')
def update_user_password():
    username = request.form.get('username', '').strip()
    new_password = request.form.get('new_password', '').strip()

    if not username:
        flash('缺少使用者名稱', 'error')
        return redirect(url_for('register'))
    if not new_password:
        flash('新密碼不可為空', 'error')
        return redirect(url_for('register'))

    with pgConnect.get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM users WHERE username = %s", (username,))
            if cur.fetchone() is None:
                flash('使用者不存在', 'error')
                return redirect(url_for('register'))

            hashed = generate_password_hash(new_password)
            cur.execute(
                "UPDATE users SET password_hash = %s WHERE username = %s",
                (hashed, username)
            )
    flash(f'{username} 的密碼已更新', 'success')
    return redirect(url_for('register'))


# ══════════════════════════════════════════════════════════════
# 批次匯入功能（equipments / contacts 共用的核心邏輯）
# ══════════════════════════════════════════════════════════════
# 流程：
#   1. 使用者上傳 JSON 或 Excel(.xlsx) 檔案
#   2. 後端解析檔案，跟資料庫現有資料比對，產生「預覽清單」
#      （標示每筆是 新增/更新/內容相同/檔案內重複/格式錯誤）
#   3. 預覽清單暫存進 Redis（10 分鐘後自動過期），回傳一個 token
#   4. 使用者在預覽畫面勾選要套用的項目，送出後端才會真正寫入 PostgreSQL
#
# 這樣設計是為了避免使用者上傳錯誤檔案就整批覆蓋資料庫，
# 一定要先看過預覽、確認無誤才會真正動到資料。

IMPORT_PREVIEW_TTL_SEC = 600  # 預覽資料在 Redis 裡保留 10 分鐘，超過要重新上傳


def _read_uploaded_records(file_storage, outer_key_is_id):
    """把上傳的 JSON 或 Excel 檔案，統一轉成 [{欄位: 值, ...}, ...] 的清單。

    outer_key_is_id：
        equipments 的 JSON 格式是 {"EQPID": {其餘欄位...}}，EQPID 是外層 key，
        所以 outer_key_is_id=True 時，會把外層 key 補進每筆資料的 'EQPID' 欄位。
        contacts 的 JSON 格式是 {"任意序號": {"EQPID": ..., 其餘欄位...}}，
        EQPID 已經在內層欄位裡，外層 key 只是序號沒有意義，
        所以 outer_key_is_id=False 時，會直接忽略外層 key。
    """
    filename = file_storage.filename or ''
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    if ext == 'json':
        raw = file_storage.read()
        try:
            data = json.loads(raw.decode('utf-8-sig'))
        except Exception as e:
            raise ValueError(f"JSON 格式錯誤，請確認檔案內容是合法的 JSON：{e}")
        if not isinstance(data, dict):
            raise ValueError(
                "JSON 內容必須是「物件」格式（key-value），"
                '例如 {"EQPID1": {...}, "EQPID2": {...}}'
            )
        records = []
        for key, val in data.items():
            if not isinstance(val, dict):
                continue
            rec = dict(val)
            if outer_key_is_id:
                rec.setdefault('EQPID', key)
            records.append(rec)
        return records

    elif ext == 'xlsx':
        if load_workbook is None:
            raise ValueError(
                "伺服器尚未安裝 openpyxl 套件，無法解析 Excel 檔案。"
                "請改用 JSON 格式上傳，或請系統管理員在 requirements.txt 加入 "
                "openpyxl 並重新建置 Docker image。"
            )
        try:
            wb = load_workbook(io.BytesIO(file_storage.read()), data_only=True)
            ws = wb.active
        except Exception as e:
            raise ValueError(f"無法讀取 Excel 檔案：{e}")

        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            raise ValueError("Excel 檔案是空的，或缺少標題列。")
        headers = [str(h).strip() if h is not None else '' for h in headers]

        records = []
        for row in rows_iter:
            if row is None or all(v is None or str(v).strip() == '' for v in row):
                continue  # 跳過空白列
            rec = {}
            for h, v in zip(headers, row):
                if not h:
                    continue
                rec[h] = '' if v is None else str(v).strip()
            records.append(rec)
        return records

    elif ext == 'xls':
        raise ValueError("目前不支援舊版 .xls 格式，請另存成 .xlsx 或改用 JSON 格式。")
    else:
        raise ValueError("不支援的檔案格式，請上傳 .json 或 .xlsx 檔案。")


def _normalize_record_keys(rec, canonical_fields):
    """把 Excel/JSON 讀進來的欄位名稱，不分大小寫比對成標準欄位名稱，
    避免使用者輸入 eqpid / EqpId / EQPID 這種大小寫不一致的狀況而漏掉欄位。
    不在 canonical_fields 清單裡的多餘欄位會被忽略。"""
    lower_map = {c.lower(): c for c in canonical_fields}
    normalized = {}
    for k, v in rec.items():
        canon = lower_map.get(str(k).strip().lower())
        if canon:
            normalized[canon] = v
    return normalized


def build_equipments_import_preview(records):
    """比對上傳的 equipments 資料跟資料庫現況，產生預覽清單。
    識別依據：EQPID（跟 equipments 表的主鍵一致）。"""
    with pgConnect.get_conn() as conn:
        with pgConnect.dict_cursor(conn) as cur:
            cur.execute("SELECT * FROM equipments")
            existing_rows = cur.fetchall()
    existing_map = {row['eqpid']: row for row in existing_rows}

    required = ['EQPTYPE', 'TESTERIP', 'PROBERIP', 'LINEGROUP', 'floor', 'Action']
    canonical = ['EQPID'] + required

    seen_in_file = set()
    preview = []
    counts = {'new': 0, 'update': 0, 'no_change': 0, 'error': 0, 'duplicate_in_file': 0}

    for raw in records:
        rec = _normalize_record_keys(raw, canonical)
        eqpid = str(rec.get('EQPID', '')).strip()
        item = {'eqpid': eqpid, 'data': rec}

        if not eqpid:
            item['status'] = 'error'
            item['reason'] = '缺少 EQPID'
            counts['error'] += 1
            preview.append(item)
            continue

        if eqpid in seen_in_file:
            item['status'] = 'duplicate_in_file'
            item['reason'] = '檔案內重複的 EQPID（已忽略，僅套用第一筆出現的資料）'
            counts['duplicate_in_file'] += 1
            preview.append(item)
            continue
        seen_in_file.add(eqpid)

        missing = [f for f in required if not str(rec.get(f, '')).strip()]
        if missing:
            item['status'] = 'error'
            item['reason'] = f"缺少欄位：{', '.join(missing)}"
            counts['error'] += 1
            preview.append(item)
            continue

        existing = existing_map.get(eqpid)
        if existing:
            old = {
                'EQPTYPE': existing['eqptype'] or '',
                'TESTERIP': existing['testerip'] or '',
                'PROBERIP': existing['proberip'] or '',
                'LINEGROUP': existing['linegroup'] or '',
                'floor': existing['floor'] or '',
                'Action': existing['action'] or '',
            }
            new_vals = {f: str(rec.get(f, '')).strip() for f in required}
            if new_vals == old:
                item['status'] = 'no_change'
                counts['no_change'] += 1
            else:
                item['status'] = 'update'
                item['old'] = old
                counts['update'] += 1
        else:
            item['status'] = 'new'
            counts['new'] += 1

        preview.append(item)

    return preview, counts


def build_contacts_import_preview(records):
    """比對上傳的 contacts 資料跟資料庫現況，產生預覽清單。
    contacts 沒有天然唯一鍵，識別依據改用 (EQPID, Action) 組合
    ——同一設備、同一種觸發動作視為同一筆聯絡對應。"""
    with pgConnect.get_conn() as conn:
        with pgConnect.dict_cursor(conn) as cur:
            cur.execute("SELECT * FROM contacts")
            existing_rows = cur.fetchall()
    existing_map = {}
    for row in existing_rows:
        key = ((row['eqpid'] or '').strip(), (row['action'] or '').strip())
        existing_map[key] = row

    required = ['EQPID', 'EQPTYPE', 'Action', 'LINEGROUP', 'floor']

    seen_in_file = set()
    preview = []
    counts = {'new': 0, 'update': 0, 'no_change': 0, 'error': 0, 'duplicate_in_file': 0}

    for raw in records:
        rec = _normalize_record_keys(raw, required)
        item = {'data': rec}

        missing = [f for f in required if not str(rec.get(f, '')).strip()]
        if missing:
            item['status'] = 'error'
            item['reason'] = f"缺少欄位：{', '.join(missing)}"
            counts['error'] += 1
            preview.append(item)
            continue

        key = (str(rec['EQPID']).strip(), str(rec['Action']).strip())
        item['eqpid'] = key[0]

        if key in seen_in_file:
            item['status'] = 'duplicate_in_file'
            item['reason'] = '檔案內重複（相同 EQPID + Action，已忽略，僅套用第一筆）'
            counts['duplicate_in_file'] += 1
            preview.append(item)
            continue
        seen_in_file.add(key)

        existing = existing_map.get(key)
        if existing:
            old = {
                'EQPID': existing['eqpid'] or '',
                'EQPTYPE': existing['eqptype'] or '',
                'Action': existing['action'] or '',
                'LINEGROUP': existing['linegroup'] or '',
                'floor': existing['floor'] or '',
            }
            new_vals = {f: str(rec.get(f, '')).strip() for f in required}
            if new_vals == old:
                item['status'] = 'no_change'
                counts['no_change'] += 1
            else:
                item['status'] = 'update'
                item['old'] = old
                item['contact_id'] = existing['id']
                counts['update'] += 1
        else:
            item['status'] = 'new'
            counts['new'] += 1

        preview.append(item)

    return preview, counts


def _store_import_preview(kind, preview):
    """把預覽清單暫存進 Redis，設定 TTL 過期，回傳一個隨機 token 供之後confirm 用。"""
    token = uuid.uuid4().hex
    key = f"import_preview:{kind}:{token}"
    redisConnect.redis_master.setex(key, IMPORT_PREVIEW_TTL_SEC, json.dumps(preview, ensure_ascii=False))
    return token


def _load_import_preview(kind, token):
    key = f"import_preview:{kind}:{token}"
    raw = redisConnect.redis_master.get(key)
    if not raw:
        return None
    return json.loads(raw)


def _delete_import_preview(kind, token):
    key = f"import_preview:{kind}:{token}"
    redisConnect.redis_master.delete(key)


@app.route('/equipments/import', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def equipments_import():
    """設備批次匯入：GET 顯示上傳表單，POST 解析檔案並顯示預覽畫面。"""
    if request.method == 'GET':
        return render_template('equipments_import.html', active_page='equipments')

    file = request.files.get('import_file')
    if not file or not file.filename:
        flash('請選擇要上傳的檔案', 'error')
        return redirect(url_for('equipments_import'))

    try:
        raw_records = _read_uploaded_records(file, outer_key_is_id=True)
    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('equipments_import'))

    if not raw_records:
        flash('檔案內沒有可匯入的資料', 'error')
        return redirect(url_for('equipments_import'))

    preview, counts = build_equipments_import_preview(raw_records)
    token = _store_import_preview('equipments', preview)

    return render_template(
        'equipments_import_preview.html',
        preview=preview,
        counts=counts,
        token=token,
        active_page='equipments'
    )


@app.route('/equipments/import/confirm', methods=['POST'])
@login_required
@role_required('admin')
def equipments_import_confirm():
    """使用者在預覽畫面勾選項目後，這裡才真正寫入 PostgreSQL。"""
    token = request.form.get('token', '')
    selected_indices = request.form.getlist('selected_indices[]')
    current_user = session.get('username', '')

    preview = _load_import_preview('equipments', token)
    if preview is None:
        flash('匯入逾時或資料已失效（預覽保留 10 分鐘），請重新上傳檔案。', 'error')
        return redirect(url_for('equipments_import'))

    created = updated = 0
    try:
        with pgConnect.get_conn() as conn:
            # 讓 equipment_history 的 trigger 記得是誰做的匯入異動
            pgConnect.set_current_user(conn, current_user)
            with conn.cursor() as cur:
                for idx_str in selected_indices:
                    try:
                        idx = int(idx_str)
                    except ValueError:
                        continue
                    if idx < 0 or idx >= len(preview):
                        continue
                    item = preview[idx]
                    if item['status'] not in ('new', 'update'):
                        continue
                    rec = item['data']
                    eqpid = item['eqpid']

                    if item['status'] == 'new':
                        cur.execute(
                            """
                            INSERT INTO equipments (eqpid, eqptype, testerip, proberip, linegroup, floor, action)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                eqpid,
                                rec.get('EQPTYPE', ''), rec.get('TESTERIP', ''), rec.get('PROBERIP', ''),
                                rec.get('LINEGROUP', ''), rec.get('floor', ''), rec.get('Action', ''),
                            )
                        )
                        created += 1
                    else:
                        cur.execute(
                            """
                            UPDATE equipments
                               SET eqptype = %s, testerip = %s, proberip = %s,
                                   linegroup = %s, floor = %s, action = %s,
                                   updated_at = now()
                             WHERE eqpid = %s
                            """,
                            (
                                rec.get('EQPTYPE', ''), rec.get('TESTERIP', ''), rec.get('PROBERIP', ''),
                                rec.get('LINEGROUP', ''), rec.get('floor', ''), rec.get('Action', ''), eqpid,
                            )
                        )
                        updated += 1
        _delete_import_preview('equipments', token)
        flash(f'匯入完成：新增 {created} 筆、更新 {updated} 筆。', 'success')
    except Exception as e:
        flash(f'匯入失敗：{e}', 'error')

    return redirect(url_for('equipments'))


@app.route('/equipments/import/cancel', methods=['POST'])
@login_required
@role_required('admin')
def equipments_import_cancel():
    token = request.form.get('token', '')
    _delete_import_preview('equipments', token)
    flash('已取消匯入。', 'success')
    return redirect(url_for('equipments'))


@app.route('/contacts/import', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def contacts_import():
    """聯絡人批次匯入：GET 顯示上傳表單，POST 解析檔案並顯示預覽畫面。"""
    if request.method == 'GET':
        return render_template('contacts_import.html', active_page='contacts')

    file = request.files.get('import_file')
    if not file or not file.filename:
        flash('請選擇要上傳的檔案', 'error')
        return redirect(url_for('contacts_import'))

    try:
        raw_records = _read_uploaded_records(file, outer_key_is_id=False)
    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('contacts_import'))

    if not raw_records:
        flash('檔案內沒有可匯入的資料', 'error')
        return redirect(url_for('contacts_import'))

    preview, counts = build_contacts_import_preview(raw_records)
    token = _store_import_preview('contacts', preview)

    return render_template(
        'contacts_import_preview.html',
        preview=preview,
        counts=counts,
        token=token,
        active_page='contacts'
    )


@app.route('/contacts/import/confirm', methods=['POST'])
@login_required
@role_required('admin')
def contacts_import_confirm():
    """使用者在預覽畫面勾選項目後，這裡才真正寫入 PostgreSQL。"""
    token = request.form.get('token', '')
    selected_indices = request.form.getlist('selected_indices[]')
    current_user = session.get('username', '')

    preview = _load_import_preview('contacts', token)
    if preview is None:
        flash('匯入逾時或資料已失效（預覽保留 10 分鐘），請重新上傳檔案。', 'error')
        return redirect(url_for('contacts_import'))

    created = updated = 0
    try:
        with pgConnect.get_conn() as conn:
            # 讓 contacts_history 的 trigger 記得是誰做的匯入異動
            pgConnect.set_current_user(conn, current_user)
            with conn.cursor() as cur:
                for idx_str in selected_indices:
                    try:
                        idx = int(idx_str)
                    except ValueError:
                        continue
                    if idx < 0 or idx >= len(preview):
                        continue
                    item = preview[idx]
                    if item['status'] not in ('new', 'update'):
                        continue
                    rec = item['data']

                    if item['status'] == 'new':
                        cur.execute(
                            """
                            INSERT INTO contacts (eqpid, eqptype, action, linegroup, floor)
                            VALUES (%s, %s, %s, %s, %s)
                            """,
                            (
                                rec.get('EQPID', ''), rec.get('EQPTYPE', ''), rec.get('Action', ''),
                                rec.get('LINEGROUP', ''), rec.get('floor', ''),
                            )
                        )
                        created += 1
                    else:
                        contact_id = item.get('contact_id')
                        cur.execute(
                            """
                            UPDATE contacts
                               SET eqpid = %s, eqptype = %s, action = %s,
                                   linegroup = %s, floor = %s
                             WHERE id = %s
                            """,
                            (
                                rec.get('EQPID', ''), rec.get('EQPTYPE', ''), rec.get('Action', ''),
                                rec.get('LINEGROUP', ''), rec.get('floor', ''), contact_id,
                            )
                        )
                        updated += 1
        _delete_import_preview('contacts', token)
        flash(f'匯入完成：新增 {created} 筆、更新 {updated} 筆。', 'success')
    except Exception as e:
        flash(f'匯入失敗：{e}', 'error')

    return redirect(url_for('contacts'))


@app.route('/contacts/import/cancel', methods=['POST'])
@login_required
@role_required('admin')
def contacts_import_cancel():
    token = request.form.get('token', '')
    _delete_import_preview('contacts', token)
    flash('已取消匯入。', 'success')
    return redirect(url_for('contacts'))


@app.route('/equipments', methods=['GET', 'POST'])
def equipments():
    error = None
    # hash_data 維持跟 Jinja 樣板相同的結構：{eqpid: json字串}，
    # 這樣 equipments.html 裡原本的 |from_json 樣板邏輯不用改
    hash_data = {}
    user_role = session.get('role', 'viewer')

    # Get filter criteria from request args, converting to uppercase for case-insensitive matching
    filter_floor = request.args.get('floor', '').strip()
    filter_eqptype = request.args.get('eqptype', '').strip()
    filter_eqpid = request.args.get('eqpid', '').strip()

    # ── 分頁參數（跟 Queue 內容查詢 /index 的邏輯一致）──────────────
    # per_page 除了 10/25/50/100 之外，多支援 'all' 代表不分頁、一次顯示全部。
    try:
        page = int(request.args.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    if page < 1:
        page = 1

    per_page_raw = request.args.get('per_page', '10').strip()
    if per_page_raw == 'all':
        per_page = 'all'
    else:
        try:
            per_page = int(per_page_raw)
        except (ValueError, TypeError):
            per_page = 10
        if per_page not in [10, 25, 50, 100]:
            per_page = 10

    total_count = 0
    total_pages = 1

    try:
        conditions = []
        params = []
        if filter_floor:
            conditions.append("floor ILIKE %s")
            params.append(filter_floor)
        if filter_eqptype:
            conditions.append("eqptype ILIKE %s")
            params.append(filter_eqptype)
        if filter_eqpid:
            conditions.append("eqpid ILIKE %s")
            params.append(filter_eqpid)

        where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

        with pgConnect.get_conn() as conn:
            with pgConnect.dict_cursor(conn) as cur:
                # 先算符合篩選條件的總筆數，才能算出總頁數／驗證目前頁碼是否超出範圍
                cur.execute(f"SELECT count(*) AS cnt FROM equipments {where_clause}", params)
                total_count = cur.fetchone()['cnt']

                if per_page == 'all':
                    total_pages = 1
                    page = 1
                    cur.execute(f"SELECT * FROM equipments {where_clause} ORDER BY eqpid", params)
                else:
                    total_pages = (total_count + per_page - 1) // per_page if total_count else 1
                    if page > total_pages:
                        page = total_pages
                    offset = (page - 1) * per_page
                    cur.execute(
                        f"SELECT * FROM equipments {where_clause} ORDER BY eqpid LIMIT %s OFFSET %s",
                        params + [per_page, offset]
                    )
                rows = cur.fetchall()

        for row in rows:
            hash_data[row['eqpid']] = json.dumps({
                'EQPTYPE': row['eqptype'],
                'TESTERIP': row['testerip'],
                'PROBERIP': row['proberip'],
                'LINEGROUP': row['linegroup'],
                'floor': row['floor'],
                'Action': row['action'],
            }, ensure_ascii=False)

        if not hash_data and (filter_floor or filter_eqptype or filter_eqpid):
            error = "沒有找到符合篩選條件的資料。"
        elif not hash_data:
            error = "「equipments」沒有內容。"

    except Exception as e:
        error = "無法連接到資料庫，請稍後再試。"
        hash_data = {}

    return render_template(
        'equipments.html',
        error=error,
        hash_data=hash_data,
        user_role=user_role,
        filter_floor=filter_floor,
        filter_eqptype=filter_eqptype,
        filter_eqpid=filter_eqpid,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        total_count=total_count,
        active_page='equipments'
    )

def is_valid_json(s):
    try:
        json.loads(s)
        return True
    except:
        return False
    
@app.route('/script_update', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def script_update():
    import requests
    import json

    # 讀 conf.json
    CONFIG_PATH = os.path.join(os.path.dirname(__file__), "conf.json")
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        conf_list = json.load(f)

    # 從 worker_status_partial 抓所有 RPA
    filtered_workers = {}
    queues = ["prober_worker_status", "LineNotify_worker_status", "LotActions_worker_status"]
    try:
        for queue_name in queues:
            all_workers = redisConnect.redis_master.hgetall(queue_name)
            for k, v in all_workers.items():
                key_str = k.decode() if isinstance(k, bytes) else str(k)
                val_str = v.decode() if isinstance(v, bytes) else str(v)
                try:
                    json_val = json.loads(val_str)
                    status = json_val.get("status", "unknown").strip().lower()
                except Exception:
                    status = "unknown"
                filtered_workers[key_str] = status
    except redis.RedisError:
        filtered_workers = {}

    rpa_list = [{"name": name, "status": status} for name, status in filtered_workers.items()]

    if request.method == 'POST':
        selected_rpas = request.form.getlist('rpa')
        source_path = request.form.get('source_path', '').strip()
        target_path = request.form.get('target_path', '').strip()

        if not selected_rpas:
            flash("請至少選擇一個 RPA", "error")
            return redirect(url_for('script_update'))
        if not source_path or not target_path:
            flash("Source Path 與 Target Path 為必填欄位", "error")
            return redirect(url_for('script_update'))

        for rpa_name in selected_rpas:
            conf = next((c for c in conf_list if c['name'] == rpa_name), None)
            if not conf:
                flash(f"{rpa_name} 對應的設定不存在", "error")
                continue

            payload = {
                "projectID": "86HklPenDDAsPmtc",
                "taskID": "KT59-m35TfGliLDH",
                "args" : [source_path,target_path]
            }

            api_url = conf.get('url')
            headers = {"Content-Type": "application/json", "X-API-KEY": conf.get("X-API-KEY")}

            try:
                response = requests.post(api_url, json=payload, headers=headers, timeout=15,verify=False )
                if response.status_code == 200:
                    flash(f"{rpa_name} 指令已成功發送", "success")
                else:
                    flash(f"{rpa_name} API 回傳錯誤：{response.status_code},{response}", "error")
            except Exception as e:
                flash(f"{rpa_name} 發送 API 失敗：{e}", "error")

        return redirect(url_for('script_update'))

    return render_template('script_update.html', rpa_list=rpa_list, active_page='script_update')

@app.route('/update_contacts', methods=['POST'])
@login_required
def update_contacts():
    edit_mode = request.form.get('edit_mode', '0')
    current_user = session.get('username', '')

    if edit_mode == '1':
        with pgConnect.get_conn() as conn:
            # 讓 contacts_history 的 trigger 記得是誰做的異動
            pgConnect.set_current_user(conn, current_user)

            with conn.cursor() as cur:
                cur.execute("SELECT id, eqpid, eqptype, action, linegroup, floor FROM contacts")
                existing_map = {
                    row[0]: (row[1] or '', row[2] or '', row[3] or '', row[4] or '', row[5] or '')
                    for row in cur.fetchall()
                }
                existing_ids = list(existing_map.keys())

                for key in existing_ids:
                    if request.form.get(f'delete_{key}'):
                        cur.execute("DELETE FROM contacts WHERE id = %s", (key,))
                        continue

                    eqpid = request.form.get(f'EQPID_{key}', '').strip()
                    eqptype = request.form.get(f'EQPTYPE_{key}', '').strip()
                    action = request.form.get(f'ACTION_{key}', '').strip()
                    linegroup = request.form.get(f'LINEGROUP_{key}', '').strip()
                    floor = request.form.get(f'floor_{key}', '').strip()

                    if eqpid and eqptype and action and linegroup and floor:
                        new_vals = (eqpid, eqptype, action, linegroup, floor)
                        old_vals = existing_map.get(key)
                        if old_vals != new_vals:
                            cur.execute(
                                """
                                UPDATE contacts
                                   SET eqpid = %s, eqptype = %s, action = %s,
                                       linegroup = %s, floor = %s
                                 WHERE id = %s
                                """,
                                (eqpid, eqptype, action, linegroup, floor, key)
                            )

                new_eqpids = request.form.getlist('new_EQPID[]')
                new_eqptypes = request.form.getlist('new_EQPTYPE[]')
                new_actions = request.form.getlist('new_ACTION[]')
                new_linegroups = request.form.getlist('new_LINEGROUP[]')
                new_floors = request.form.getlist('new_floor[]')

                new_row_count = max(
                    [len(new_eqpids), len(new_eqptypes), len(new_actions), len(new_linegroups), len(new_floors)],
                    default=0
                )
                for i in range(new_row_count):
                    eqpid = new_eqpids[i].strip() if i < len(new_eqpids) else ''
                    eqptype = new_eqptypes[i].strip() if i < len(new_eqptypes) else ''
                    action = new_actions[i].strip() if i < len(new_actions) else ''
                    linegroup = new_linegroups[i].strip() if i < len(new_linegroups) else ''
                    floor = new_floors[i].strip() if i < len(new_floors) else ''

                    if all(not value for value in [eqpid, floor, eqptype, action, linegroup]):
                        continue

                    # id 由 PostgreSQL 的 SERIAL 自動產生，不用再手動 INCR
                    cur.execute(
                        """
                        INSERT INTO contacts (eqpid, eqptype, action, linegroup, floor)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (eqpid, eqptype, action, linegroup, floor)
                    )

        return redirect(url_for("contacts"))

    return redirect(url_for("contacts"))


@app.route('/contacts', methods=['GET', 'POST'])
def contacts():
    error = None
    # hash_data 維持跟原本一樣的 {id: json字串} 結構，相容原本的 contacts.html 樣板
    hash_data = {}
    user_role = session.get('role', 'viewer')

    filter_floor = request.args.get('floor', '').strip()
    filter_eqptype = request.args.get('eqptype', '').strip()
    filter_eqpid = request.args.get('eqpid', '').strip()

    # ── 分頁參數（跟 /equipments 的邏輯一致）──────────────────────
    try:
        page = int(request.args.get('page', 1))
    except (ValueError, TypeError):
        page = 1
    if page < 1:
        page = 1

    per_page_raw = request.args.get('per_page', '10').strip()
    if per_page_raw == 'all':
        per_page = 'all'
    else:
        try:
            per_page = int(per_page_raw)
        except (ValueError, TypeError):
            per_page = 10
        if per_page not in [10, 25, 50, 100]:
            per_page = 10

    total_count = 0
    total_pages = 1

    try:
        conditions = []
        params = []
        if filter_floor:
            conditions.append("floor ILIKE %s")
            params.append(filter_floor)
        if filter_eqptype:
            conditions.append("eqptype ILIKE %s")
            params.append(filter_eqptype)
        if filter_eqpid:
            conditions.append("eqpid ILIKE %s")
            params.append(filter_eqpid)

        where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

        with pgConnect.get_conn() as conn:
            with pgConnect.dict_cursor(conn) as cur:
                cur.execute(f"SELECT count(*) AS cnt FROM contacts {where_clause}", params)
                total_count = cur.fetchone()['cnt']

                if per_page == 'all':
                    total_pages = 1
                    page = 1
                    cur.execute(f"SELECT * FROM contacts {where_clause} ORDER BY id", params)
                else:
                    total_pages = (total_count + per_page - 1) // per_page if total_count else 1
                    if page > total_pages:
                        page = total_pages
                    offset = (page - 1) * per_page
                    cur.execute(
                        f"SELECT * FROM contacts {where_clause} ORDER BY id LIMIT %s OFFSET %s",
                        params + [per_page, offset]
                    )
                rows = cur.fetchall()

        for row in rows:
            hash_data[str(row['id'])] = json.dumps({
                'EQPID': row['eqpid'],
                'EQPTYPE': row['eqptype'],
                'Action': row['action'],
                'LINEGROUP': row['linegroup'],
                'floor': row['floor'],
            }, ensure_ascii=False)

        if not hash_data and (filter_floor or filter_eqptype or filter_eqpid):
            error = "沒有找到符合篩選條件的資料。"
        elif not hash_data:
            error = "「contacts」沒有內容。"

    except Exception as e:
        error = "無法連接到資料庫，請稍後再試。"
        hash_data = {}

    return render_template(
        'contacts.html',
        error=error,
        hash_data=hash_data,
        user_role=user_role,
        filter_floor=filter_floor,
        filter_eqptype=filter_eqptype,
        filter_eqpid=filter_eqpid,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        total_count=total_count,
        active_page='contacts'
    )


def load_contact_history(eqpid=''):
    """查詢聯絡人異動紀錄（contacts_history 表由 PostgreSQL trigger 自動寫入，
    見 init_contacts_history_schema()）。eqpid 有值時只查該 EQPID 相關的紀錄，
    沒有值時回傳最近 200 筆全部紀錄。"""
    history = []
    history_error = None
    try:
        with pgConnect.get_conn() as conn:
            with pgConnect.dict_cursor(conn) as cur:
                if eqpid:
                    cur.execute(
                        """
                        SELECT * FROM contacts_history
                         WHERE eqpid = %s
                         ORDER BY changed_at DESC
                         LIMIT 200
                        """,
                        (eqpid,)
                    )
                else:
                    cur.execute(
                        """
                        SELECT * FROM contacts_history
                         ORDER BY changed_at DESC
                         LIMIT 200
                        """
                    )
                history = cur.fetchall()
        if not history:
            history_error = "沒有找到異動紀錄。"
    except Exception as e:
        history_error = f"查詢失敗：{e}"
    return history, history_error


@app.route('/contact_history')
@login_required
def contact_history():
    """查看聯絡人異動紀錄（獨立頁面，進入按鈕放在 /contacts 頁面裡）。"""
    eqpid = request.args.get('eqpid', '').strip()
    history, error = load_contact_history(eqpid)
    return render_template(
        'contact_history.html',
        history=history,
        eqpid=eqpid,
        error=error,
        active_page='contact_history'
    )

def compute_worker_utilization(queue_name, cutoff, end_dt, interval_min, worker_filter=''):
    """
    計算某個 worker_status Queue（LotActions_worker_status / prober_worker_status /
    LineNotify_worker_status）底下，每個 RPA worker 在 [cutoff, end_dt] 區間內的稼動率（%）。

    資料來源：
      worker_status_history 表，由本檔案啟動時開啟的背景執行緒（worker_status_sync_loop）
      持續輪詢 Redis 的 worker_status Hash，偵測到 idle/busy/offline 狀態變化時才寫入
      一筆紀錄（狀態沒變化不會重複寫入）。這是目前最準確的稼動率計算方式，
      直接反映 worker_status 的真實 busy/idle 切換，不需要靠 EQPID 或
      LotStart/LotEnd 事件去推算操作區間。

    計算邏輯：
      稼動率 = 該 worker 在時間桶內處於 'busy' 狀態的秒數 / 桶子總秒數

      因為狀態紀錄是「事件式」的（只在變化時寫入），所以要先重建出每個 worker
      在查詢區間內的完整狀態時間軸：
        1. 查出每個 worker 在 cutoff 當下的「起始狀態」
           （cutoff 之前最後一次變化紀錄的狀態）。
        2. 查出 (cutoff, end_dt] 區間內的所有狀態變化紀錄。
        3. 把「起始狀態」跟後續變化串起來，重建出每一段的
           開始時間、結束時間、狀態值。
        4. 如果 worker 在區間內最後一次的狀態是 busy，且之後沒有再變化
           （代表目前還在忙碌中），這段忙碌會自動延伸到 end_dt，
           確保即時畫面能反映當下正在忙碌的 worker。

    已知限制：
      這套方法只能從「這個 Flask 應用程式啟動之後」才有歷史資料可查，
      之前的 busy/idle 變化沒有被記錄下來。
    """
    local_tz = timezone(timedelta(hours=8))

    initial_sql = """
        SELECT DISTINCT ON (worker_name) worker_name, status, changed_at
        FROM worker_status_history
        WHERE queue_name = %(queue_name)s AND changed_at <= %(cutoff)s
        ORDER BY worker_name, changed_at DESC
    """
    changes_sql = """
        SELECT worker_name, status, changed_at
        FROM worker_status_history
        WHERE queue_name = %(queue_name)s
          AND changed_at > %(cutoff)s AND changed_at <= %(end)s
        ORDER BY worker_name, changed_at
    """
    params = {
        'queue_name': queue_name,
        'cutoff': cutoff,
        'end': end_dt,
    }

    with pgConnect.get_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(initial_sql, params)
            initial_rows = cur.fetchall()
            cur.execute(changes_sql, params)
            change_rows = cur.fetchall()

    # 組成每個 worker 的事件時間軸：{worker_name: [(time, status), ...]}
    timelines = {}
    for row in initial_rows:
        timelines.setdefault(row['worker_name'], []).append((cutoff, row['status']))
    for row in change_rows:
        ts = row['changed_at']
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=local_tz)
        timelines.setdefault(row['worker_name'], []).append((ts, row['status']))

    if worker_filter:
        timelines = {w: evs for w, evs in timelines.items() if w == worker_filter}

    # 把每個 worker 的事件時間軸，轉成 busy 區間列表 [(start, end), ...]
    busy_intervals = {}
    for worker, events in timelines.items():
        events.sort(key=lambda e: e[0])
        intervals = []
        for i, (t, status) in enumerate(events):
            seg_start = t
            seg_end = events[i + 1][0] if i + 1 < len(events) else end_dt
            if status == 'busy' and seg_end > seg_start:
                intervals.append((seg_start, seg_end))
        busy_intervals[worker] = intervals

    # 產生時間桶（跟 queue_history 的 bucket 邏輯一致，確保兩個圖表的時間軸對得起來）
    all_labels = []
    bucket_origin = cutoff.replace(minute=(cutoff.minute // interval_min) * interval_min, second=0, microsecond=0)
    ptr = bucket_origin
    bucket_bounds = []
    while ptr <= end_dt:
        b_start = ptr
        b_end = min(ptr + timedelta(minutes=interval_min), end_dt)
        all_labels.append(ptr.strftime('%m-%d %H:%M'))
        bucket_bounds.append((b_start, b_end))
        ptr += timedelta(minutes=interval_min)

    workers_set = sorted(busy_intervals.keys())
    busy_seconds = {w: [0.0] * len(bucket_bounds) for w in workers_set}

    for worker, intervals in busy_intervals.items():
        for start_ts, end_ts in intervals:
            if start_ts.tzinfo is None:
                start_ts = start_ts.replace(tzinfo=local_tz)
            if end_ts.tzinfo is None:
                end_ts = end_ts.replace(tzinfo=local_tz)
            for idx, (b_start, b_end) in enumerate(bucket_bounds):
                overlap_start = max(start_ts, b_start)
                overlap_end = min(end_ts, b_end)
                overlap = (overlap_end - overlap_start).total_seconds()
                if overlap > 0:
                    busy_seconds[worker][idx] += overlap

    series = {}
    for w in workers_set:
        pct_list = []
        for idx, (b_start, b_end) in enumerate(bucket_bounds):
            bucket_seconds = (b_end - b_start).total_seconds()
            pct = (busy_seconds[w][idx] / bucket_seconds * 100) if bucket_seconds > 0 else 0
            pct_list.append(round(min(pct, 100), 1))
        series[w] = pct_list

    all_values = [v for vals in series.values() for v in vals]
    avg_utilization = round(sum(all_values) / len(all_values), 1) if all_values else 0.0

    return {
        'labels': all_labels,
        'series': series,
        'avg_utilization': avg_utilization,
        'worker_count': len(workers_set),
    }


@app.route('/worker_utilization')
@login_required
def worker_utilization():
    """RPA Worker 稼動率趨勢頁：資料來源為 worker_status_history
    （由本檔案內建的背景執行緒 worker_status_sync_loop 持續同步 idle/busy/offline 狀態變化）。
    畫面風格與資料流跟 /queue_history 一致。"""
    available_categories = ['LotActions', 'prober', 'LineNotify']
    category = request.args.get('category', '').strip() or available_categories[0]
    # worker_status Hash 名稱跟 category 的對應規則：{category}_worker_status
    queue_name = f"{category}_worker_status"
    worker_name_filter = request.args.get('worker_name', '').strip()

    all_rpa_names = set()
    worker_queues = ["prober_worker_status", "LineNotify_worker_status", "LotActions_worker_status"]
    try:
        for q in worker_queues:
            keys = redisConnect.redis_master.hkeys(q)
            for k in keys:
                all_rpa_names.add(k)
    except Exception:
        pass
    sorted_workers = sorted(all_rpa_names)

    mode = request.args.get('mode', 'last')
    local_tz = timezone(timedelta(hours=8))
    now_dt = datetime.now(tz=local_tz)

    if mode == 'range':
        parsed_start = safe_parse_datetime(request.args.get('start_time', '').strip())
        parsed_end = safe_parse_datetime(request.args.get('end_time', '').strip())
        cutoff = parsed_start.replace(tzinfo=local_tz) if parsed_start else now_dt - timedelta(hours=24)
        end_dt = parsed_end.replace(tzinfo=local_tz) if parsed_end else now_dt
        hours = int((end_dt - cutoff).total_seconds() / 3600)
    else:
        try:
            hours = int(request.args.get('hours', 24))
        except ValueError:
            hours = 24
        if hours not in [1, 3, 6, 12, 24, 48, 72]:
            hours = 24
        cutoff = now_dt - timedelta(hours=hours)
        end_dt = now_dt

    try:
        interval_min = int(request.args.get('interval', 60))
    except ValueError:
        interval_min = 60
    if interval_min not in [10, 30, 60]:
        interval_min = 60

    chart_data = None
    error = None
    avg_utilization = 0.0
    worker_count = 0

    try:
        result = compute_worker_utilization(queue_name, cutoff, end_dt, interval_min, worker_name_filter)
        chart_data = {
            'labels': result['labels'],
            'series': result['series'],
        }
        avg_utilization = result['avg_utilization']
        worker_count = result['worker_count']
    except Exception as e:
        error = f"查詢失敗：{str(e)}"

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if error:
            return jsonify({'error': error}), 400
        return jsonify({
            'chart_data': chart_data,
            'avg_utilization': avg_utilization,
            'worker_count': worker_count,
            'category': category,
        })

    return render_template(
        'worker_utilization.html',
        category=category,
        worker_name_filter=worker_name_filter,
        sorted_workers=sorted_workers,
        hours=hours,
        mode=mode,
        start_time=request.args.get('start_time', ''),
        end_time=request.args.get('end_time', ''),
        interval=interval_min,
        chart_data=chart_data,
        error=error,
        avg_utilization=avg_utilization,
        worker_count=worker_count,
        available_categories=available_categories,
        active_page='worker_utilization',
    )


def compute_task_history_chart(category, cutoff, end_dt, interval_min, worker_filter='',
                                py_label_format='%m-%d %H:%M', sql_label_format='MM-DD HH24:MI'):
    """
    查詢 PostgreSQL task_history 表，算出指定 category 在 [cutoff, end_dt]
    區間內、每個時間桶的 dispatched/failed 筆數，回傳 chart_data 格式。

    這是從 /queue_history 抽出來的共用邏輯，讓 dashboard_data() 的趨勢圖
    也能呼叫同一套查詢，確保兩個頁面看到的趨勢圖數字完全一致，
    不再各自維護一份、容易對不起來。

    用 date_bin() 直接在資料庫端做時間分桶聚合，不需要把整個 Redis list
    撈進 Python 手動算——這是目前系統的標準做法（跟 /queue_history 一致）。

    py_label_format / sql_label_format：
        兩邊頁面原本的 X 軸標籤格式不一樣——/queue_history 顯示「月-日 時:分」
        （因為可以查好幾天的區間），/dashboard 只顯示「時:分」（因為固定只看
        最近 6 小時，不需要日期）。這兩個參數必須保持對應（同樣的日期/時間
        格式，只是 Python strftime 跟 PostgreSQL to_char 的語法不同），
        才能讓 Python 端產生的 all_labels 跟 SQL 查出來的 bucket 字串對得上。
    """
    all_labels = []
    bucket_origin = cutoff.replace(minute=(cutoff.minute // interval_min) * interval_min, second=0, microsecond=0)
    ptr = bucket_origin
    while ptr <= end_dt:
        all_labels.append(ptr.strftime(py_label_format))
        ptr += timedelta(minutes=interval_min)

    sql = f"""
        SELECT
            to_char(
                date_bin(%(interval)s, dispatch_time, %(origin)s) AT TIME ZONE 'Asia/Taipei',
                '{sql_label_format}'
            ) AS bucket,
            status,
            count(*) AS cnt
        FROM task_history
        WHERE category = %(category)s
          AND dispatch_time BETWEEN %(start)s AND %(end)s
          AND (%(worker)s = '' OR rpa_worker = %(worker)s)
        GROUP BY bucket, status
    """
    params = {
        'interval': timedelta(minutes=interval_min),
        'origin': bucket_origin,
        'category': category,
        'start': cutoff,
        'end': end_dt,
        'worker': worker_filter,
    }

    disp_bucket = defaultdict(int)
    fail_bucket = defaultdict(int)
    total_dispatched = 0
    total_failed = 0
    with pgConnect.get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            for bucket_label, status, cnt in cur.fetchall():
                if status == 'dispatched':
                    disp_bucket[bucket_label] += cnt
                    total_dispatched += cnt
                elif status == 'failed':
                    fail_bucket[bucket_label] += cnt
                    total_failed += cnt

    chart_data = {
        'labels': all_labels,
        'counts_dispatched': [disp_bucket.get(lbl, 0) for lbl in all_labels],
        'counts_failed': [fail_bucket.get(lbl, 0) for lbl in all_labels],
    }
    return chart_data, total_dispatched, total_failed


@app.route('/queue_history')
@login_required
def queue_history():
    # 限定任務類別
    available_categories = ['LotActions', 'prober', 'LineNotify']

    category = request.args.get('category', '').strip()
    if not category:
        category = available_categories[0]

    # 時間一律固定為 dispatch_time
    time_field = 'dispatch_time'
    worker_name_filter = request.args.get('worker_name', '').strip()

    # 獲取所有 RPA 清單 (worker_status 維持在 Redis，不受這次改動影響)
    all_rpa_names = set()
    worker_queues = ["prober_worker_status", "LineNotify_worker_status", "LotActions_worker_status"]
    try:
        for q in worker_queues:
            keys = redisConnect.redis_master.hkeys(q)
            for k in keys:
                all_rpa_names.add(k)
    except: pass
    sorted_workers = sorted(list(all_rpa_names))

    # 時間模式：last (最近) 或 range (區間)
    mode = request.args.get('mode', 'last')

    local_tz = timezone(timedelta(hours=8))
    now_dt = datetime.now(tz=local_tz)

    if mode == 'range':
        start_str = request.args.get('start_time', '').strip()
        end_str = request.args.get('end_time', '').strip()
        try:
            parsed_start = safe_parse_datetime(start_str)
            cutoff = parsed_start.replace(tzinfo=local_tz) if parsed_start else now_dt - timedelta(hours=24)

            parsed_end = safe_parse_datetime(end_str)
            end_dt = parsed_end.replace(tzinfo=local_tz) if parsed_end else now_dt
        except:
            cutoff = now_dt - timedelta(hours=24)
            end_dt = now_dt
        hours = int((end_dt - cutoff).total_seconds() / 3600)
    else:
        try:
            hours = int(request.args.get('hours', 24))
        except ValueError:
            hours = 24
        if hours not in [1, 3, 6, 12, 24, 48, 72]:
            hours = 24
        cutoff = now_dt - timedelta(hours=hours)
        end_dt = now_dt

    try:
        interval_min = int(request.args.get('interval', 60))
    except ValueError:
        interval_min = 60
    if interval_min not in [10, 30, 60]:
        interval_min = 60

    chart_data = None
    error = None
    total_dispatched = 0
    total_failed = 0

    if category:
        try:
            chart_data, total_dispatched, total_failed = compute_task_history_chart(
                category, cutoff, end_dt, interval_min, worker_name_filter
            )
        except Exception as e:
            error = f"查詢失敗：{str(e)}"

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if error: return jsonify({'error': error}), 400
        return jsonify({
            'chart_data': chart_data,
            'total_dispatched': total_dispatched,
            'total_failed': total_failed,
            'category': category
        })

    return render_template(
        'queue_history.html',
        category=category,
        time_field=time_field,
        worker_name_filter=worker_name_filter,
        sorted_workers=sorted_workers,
        hours=hours,
        mode=mode,
        start_time=request.args.get('start_time', ''),
        end_time=request.args.get('end_time', ''),
        interval=interval_min,
        chart_data=chart_data,
        error=error,
        total_dispatched=total_dispatched,
        total_failed=total_failed,
        available_categories=available_categories,
        active_page='queue_history'
    )

@app.route('/api_trigger', methods=['GET', 'POST'])
@login_required
def api_trigger():
    EXTERNAL_API_BASE = "http://10.97.210.35:8000"
    user_role = session.get('role', 'viewer')
    
    # 獲取所有機台編號供下拉選單使用（equipments 已搬到 PostgreSQL）
    eqp_list = []
    try:
        with pgConnect.get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT eqpid FROM equipments ORDER BY eqpid")
                eqp_list = [row[0] for row in cur.fetchall()]
    except: pass

    if request.method == 'POST':
        endpoint = request.form.get('endpoint', '/EQPAction')
        eqpid = request.form.get('eqpid', '').strip()
        action = request.form.get('action', 'LotEnd').strip()
        
        payload = { "EQPID": eqpid, "Action": action }
        
        # 針對 DataEntry 補上額外欄位
        if endpoint == '/EQPAction_DataEntry':
            payload.update({
                "TesterInput": request.form.get('tester_input', ''),
                "ProberInput": request.form.get('prober_input', ''),
                "Mode": request.form.get('mode', 'Auto')
            })

        target_url = f"{EXTERNAL_API_BASE}{endpoint}"
        try:
            if endpoint == '/ContactInquiry':
                # ContactInquiry 是 GET 請求
                response = requests.get(target_url, params=payload, timeout=5)
            else:
                # 其他是 POST 請求
                response = requests.post(target_url, json=payload, timeout=5)
            
            result = {
                "status_code": response.status_code,
                "reason": response.reason,
                "content": response.text
            }
            return render_template('api_trigger.html', eqp_list=eqp_list, result=result, last_payload=payload, endpoint=endpoint, user_role=user_role, active_page='api_trigger')
        except Exception as e:
            return render_template('api_trigger.html', eqp_list=eqp_list, error=f"連線失敗: {str(e)}", user_role=user_role, active_page='api_trigger')

    return render_template('api_trigger.html', eqp_list=eqp_list, user_role=user_role, active_page='api_trigger')

# ── Dashboard ──────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    user_role = session.get('role', 'viewer')
    return render_template('dashboard.html', user_role=user_role, active_page='dashboard')


@app.route('/dashboard_data')
@login_required
def dashboard_data():
    """Dashboard 用的單一聚合 API，前端每 5 秒輪詢一次"""
    local_tz = timezone(timedelta(hours=8))
    now_dt   = datetime.now(tz=local_tz)
    mode = request.args.get('mode', 'last')
    if mode == 'range':
        parsed_start = safe_parse_datetime(request.args.get('start_time', '').strip())
        parsed_end = safe_parse_datetime(request.args.get('end_time', '').strip())
        cutoff = parsed_start.replace(tzinfo=local_tz) if parsed_start else now_dt - timedelta(hours=6)
        end_dt = parsed_end.replace(tzinfo=local_tz) if parsed_end else now_dt
    else:
        try:
            hours = int(request.args.get('hours', 6))
        except ValueError:
            hours = 6
        cutoff = now_dt - timedelta(hours=hours)
        end_dt = now_dt

    try:
        interval_min = int(request.args.get('interval', 30))
    except ValueError:
        interval_min = 30
    if interval_min not in [10, 30, 60]:
        interval_min = 30

    cutoff   = now_dt - timedelta(hours=6)   # 趨勢圖固定顯示最近 6 小時
    interval_min = 30                         # 每 30 分鐘一個資料點

    # ── 1. Queue 長度 ──────────────────────────────────────────
    # Dashboard 面板只顯示 DASHBOARD_MONITORED_QUEUES 這 8 個重點 Queue
    # （畫面空間有限），但長度計算呼叫共用的 get_queue_length()，
    # 跟 /queue_lengths 的計算邏輯完全一致
    # （例如 LotActions_dispatched_log 會改查 PostgreSQL 累計總筆數）。
    queue_stats = []
    # 只取 ALLOWED_QUEUES 裡有的
    for q in DASHBOARD_MONITORED_QUEUES:
        if q not in ALLOWED_QUEUES:
            continue
        _, length = get_queue_length(q)
        queue_stats.append({"name": q, "length": length})

    # ── 2. Worker 狀態 ─────────────────────────────────────────
    worker_status_queues = [
        "prober_worker_status",
        "LineNotify_worker_status",
        "LotActions_worker_status",
    ]
    workers = {}
    idle_count = busy_count = ghost_count = offline_count = 0
    try:
        for wq in worker_status_queues:
            all_w = redisConnect.redis_master.hgetall(wq)
            for k, v in all_w.items():
                k = k.decode() if isinstance(k, bytes) else k
                v = v.decode() if isinstance(v, bytes) else v
                try:
                    data   = json.loads(v)
                    status = data.get("status", "unknown").strip().lower()
                except Exception:
                    status = "unknown"
                workers[k] = status
                if status == "idle":    idle_count    += 1
                elif status == "busy":  busy_count    += 1
                elif status == "ghost": ghost_count   += 1
                else:                   offline_count += 1
    except Exception:
        pass

    # ── 3. 今日任務統計（從 dispatched_log + failed_queue 計算）──
    categories = ["LotActions", "LineNotify", "prober"]
    today_start = now_dt.replace(hour=0, minute=0, second=0, microsecond=0)

    total_dispatched_today = 0
    total_failed_today     = 0

    for cat in categories:
        for suffix, is_fail in [("dispatched_log", False), ("failed_queue", True)]:
            keys = [f"queue_history:{cat}_{suffix}", f"{cat}_{suffix}"]
            raw_items = []
            for k in keys:
                try:
                    if redisConnect.redis_master.exists(k):
                        raw_items = redisConnect.redis_master.lrange(k, 0, -1)
                        if raw_items:
                            break
                except Exception:
                    pass
            for raw in raw_items:
                try:
                    item = json.loads(raw)
                    if item.get("task_id") == "__INIT__":
                        continue
                    ts = item.get("dispatch_time") or item.get("timestamp")
                    if not ts:
                        continue
                    dt = safe_parse_datetime(str(ts))
                    if not dt: continue
                    dt = dt.replace(tzinfo=local_tz) if dt.tzinfo is None else dt.astimezone(local_tz)
                    if dt < today_start:
                        continue
                    if is_fail:
                        total_failed_today += 1
                    else:
                        total_dispatched_today += 1
                except Exception:
                    pass

    # ── 4. 趨勢圖資料（使用者選擇的 category，預設 LotActions）──
    # 改成跟 /queue_history 一樣直接查 PostgreSQL task_history 表，不再讀 Redis。
    # 原因：history_sync 服務會持續把 Redis 的 dispatched_log/failed_queue
    # 搬到 task_history 並清空來源 list，所以讀 Redis 只能看到「還沒被搬走
    # 的殘留資料」，數字會嚴重低估。查 task_history 才是真正累計、準確的數字，
    # 也才能跟 /queue_history 頁面的數字互相對得起來。
    category = request.args.get("category", "LotActions")

    try:
        chart_data, _, _ = compute_task_history_chart(
            category, cutoff, now_dt, interval_min,
            py_label_format='%H:%M', sql_label_format='HH24:MI'
        )
    except Exception:
        # 查詢失敗時維持原本的空圖表格式，不讓整個 dashboard_data API 掛掉
        all_labels = []
        ptr = cutoff.replace(minute=(cutoff.minute // interval_min) * interval_min, second=0, microsecond=0)
        while ptr <= now_dt:
            all_labels.append(ptr.strftime('%H:%M'))
            ptr += timedelta(minutes=interval_min)
        chart_data = {
            "labels": all_labels,
            "counts_dispatched": [0] * len(all_labels),
            "counts_failed": [0] * len(all_labels),
        }

    return jsonify({
        "queue_stats":             queue_stats,
        "workers":                 workers,
        "worker_summary":          {
            "idle": idle_count, "busy": busy_count,
            "ghost": ghost_count, "offline": offline_count,
            "total": len(workers)
        },
        "total_dispatched_today":  total_dispatched_today,
        "total_failed_today":      total_failed_today,
        "chart_data":              chart_data,
        "category":                category,
        "generated_at":            now_dt.strftime('%H:%M:%S'),
    })


if __name__ == '__main__':
    threading.Thread(target=redisConnect.listen_for_failover, daemon=True)
    app.run(host='0.0.0.0', port=5000, debug=False)
    

# cd C:\Users\RPA_TEST\Downloads\a
# docker-compose up -d --build

# docker exec -it redis-master redis-cli
# KEYS *
# lrange task_queue 0 -1
# HSET LotActions_worker_status LotActions-RPA-BOT-01 '{"status": "busy", "floor": "2AF"}'

# lrange queue_history:prober_failed_queue 0 -1
